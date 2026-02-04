from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl

from normalize import normalize_measure_name


@dataclass(frozen=True)
class KickoutList:
    # group_name -> normalized_key -> display_name (as in Excel)
    groups: Dict[str, Dict[str, str]]
    # group_name -> list of duplicated normalized keys inside that column
    duplicates: Dict[str, List[str]]
    # for UI / debugging
    schema: str  # "welding" or "lead"
    sheet_title: str


class KickoutListError(Exception):
    pass


def _find_col(header_lower: List[str], name: str) -> Optional[int]:
    name = name.lower()
    for idx, h in enumerate(header_lower):
        if h == name:
            return idx
    return None


def load_kickout_list_xlsx(xlsx_path: Path, sheet_name: Optional[str] = None) -> KickoutList:
    if not xlsx_path.exists():
        raise KickoutListError(f"Kickout list not found: {xlsx_path}")

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    header = [
        str(c.value).strip() if c.value is not None else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    header_l = [h.lower() for h in header]

    # Detect schema
    upper_col = _find_col(header_l, "upper")
    lower_col = _find_col(header_l, "lower")

    anode_col = _find_col(header_l, "anode")
    cathode_col = _find_col(header_l, "cathode")
    shared_col = _find_col(header_l, "shared")

    if upper_col is not None and lower_col is not None:
        schema = "welding"
        group_cols = {"Upper": upper_col, "Lower": lower_col}
    elif anode_col is not None and cathode_col is not None and shared_col is not None:
        schema = "lead"
        group_cols = {"Anode": anode_col, "Cathode": cathode_col, "Shared": shared_col}
    else:
        raise KickoutListError(
            f"Unrecognized kickout schema in {xlsx_path.name} (sheet '{ws.title}'). "
            f"Need either headers: Upper/Lower OR Anode/Cathode/Shared."
        )

    groups: Dict[str, Dict[str, str]] = {g: {} for g in group_cols.keys()}
    dup_counts: Dict[str, Dict[str, int]] = {g: {} for g in group_cols.keys()}

    for row in ws.iter_rows(min_row=2):
        for group, col in group_cols.items():
            if col >= len(row):
                continue
            val = row[col].value
            if val is None:
                continue
            s = str(val).strip()
            if not s:
                continue
            key = normalize_measure_name(s)
            if not key:
                continue

            dup_counts[group][key] = dup_counts[group].get(key, 0) + 1
            # first seen display name wins (we want "as-is from master list")
            groups[group].setdefault(key, s)

    duplicates: Dict[str, List[str]] = {}
    for g, counts in dup_counts.items():
        duplicates[g] = sorted([k for k, c in counts.items() if c > 1])

    return KickoutList(groups=groups, duplicates=duplicates, schema=schema, sheet_title=ws.title)
