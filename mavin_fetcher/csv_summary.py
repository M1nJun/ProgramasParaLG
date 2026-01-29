from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import csv

# Regions we care about
REGIONS = ("LOWER_B_L", "LOWER_B_R", "UPPER_B_L", "UPPER_B_R")


@dataclass
class Summary:
    rows: int
    # region -> class_name -> count
    by_region: Dict[str, Dict[str, int]]
    # class_name -> count (sum across regions)
    overall: Dict[str, int]


def _inc(d: Dict[str, int], key: str, n: int = 1) -> None:
    d[key] = d.get(key, 0) + n


def _read_csv_dict_rows(path: Path) -> Iterable[dict]:
    with path.open("r", newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            yield row


def _read_xlsx_dict_rows(path: Path) -> Iterable[dict]:
    # Lazy import so users without openpyxl can still use CSV
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return

    headers = [str(h).strip() if h is not None else "" for h in header]

    for values in rows_iter:
        row = {}
        for i, v in enumerate(values):
            if i >= len(headers):
                break
            row[headers[i]] = "" if v is None else str(v)
        yield row


def _iter_rows(path: Path) -> Iterable[dict]:
    suf = path.suffix.lower()
    if suf == ".csv":
        return _read_csv_dict_rows(path)
    if suf in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _read_xlsx_dict_rows(path)
    raise ValueError(f"Unsupported file type: {path.name} (expected .csv or .xlsx)")


def summarize(paths: List[Path]) -> Summary:
    by_region: Dict[str, Dict[str, int]] = {r: {} for r in REGIONS}
    overall: Dict[str, int] = {}
    total_rows = 0

    for p in paths:
        for row in _iter_rows(p):
            total_rows += 1
            for region in REGIONS:
                col = f"{region}-NAME"
                name = (row.get(col) or "").strip()
                if not name:
                    continue
                _inc(by_region[region], name, 1)
                _inc(overall, name, 1)

    return Summary(rows=total_rows, by_region=by_region, overall=overall)


def format_summary(s: Summary, top_n: int = 20) -> str:
    lines: List[str] = []
    lines.append(f"[CSV] Total rows: {s.rows}")

    def top_items(d: Dict[str, int]) -> List[Tuple[str, int]]:
        return sorted(d.items(), key=lambda kv: (-kv[1], kv[0]))[:top_n]

    lines.append("")
    lines.append("[CSV] Overall (sum across 4 regions) - top counts:")
    for k, v in top_items(s.overall):
        lines.append(f"  - {k}: {v}")

    lines.append("")
    lines.append("[CSV] By region - top counts:")
    for region in REGIONS:
        lines.append(f"  {region}:")
        items = top_items(s.by_region.get(region, {}))
        if not items:
            lines.append("    (no data)")
        else:
            for k, v in items:
                lines.append(f"    - {k}: {v}")

    return "\n".join(lines)
