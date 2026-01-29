from __future__ import annotations

import csv
from pathlib import Path
from typing import Dict, Iterable, Iterator

from openpyxl import load_workbook


def iter_rows(path: Path) -> Iterator[Dict[str, str]]:
    """
    Yield each row as a dict[str,str] from either:
      - .csv
      - .xlsx / .xlsm
    """
    path = path.expanduser().resolve()
    suf = path.suffix.lower()

    if suf == ".csv":
        yield from _iter_csv(path)
        return

    if suf in (".xlsx", ".xlsm"):
        yield from _iter_xlsx(path)
        return

    raise ValueError(f"Unsupported file type: {path.name}")


def _iter_csv(path: Path) -> Iterator[Dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.DictReader(f)
        for row in reader:
            # ensure all values are strings
            out = {}
            for k, v in row.items():
                if k is None:
                    continue
                out[str(k)] = "" if v is None else str(v)
            yield out


def _iter_xlsx(path: Path) -> Iterator[Dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = ws.iter_rows(values_only=True)

        header = next(rows, None)
        if not header:
            return

        headers = ["" if h is None else str(h) for h in header]

        for r in rows:
            row = {}
            for i, val in enumerate(r):
                key = headers[i] if i < len(headers) else ""
                if not key:
                    continue
                row[key] = "" if val is None else str(val)
            yield row
    finally:
        wb.close()
